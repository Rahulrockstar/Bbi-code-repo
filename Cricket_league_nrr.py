import openpyxl
from decimal import Decimal

# Load the workbook
workbook = openpyxl.load_workbook('C:\\cricket_scores.xlsx')

# Get the active sheet
worksheet = workbook.active

# Check if the "Points and NRR" sheet exists
if "Points and NRR" in workbook.sheetnames:
    points_worksheet = workbook["Points and NRR"]
else:
    # If the "Points and NRR" sheet doesn't exist, create a new sheet
    points_worksheet = workbook.create_sheet("Points and NRR")

# Define column headers for the second sheet
headers = ["Team Name", "Matches Played", "Win", "Loss", "Points", "NRR"]

# Write headers to the second sheet if it's a newly created sheet
if points_worksheet.max_row == 1:
    points_worksheet.append(headers)

# Initialize variables for team statistics
teams = {
    "HBK": {
        "Total_runs_scored": 0,
        "Total_runs_conceded": 0,
        "Total_overs_played": Decimal(0),
        "Total_Matches": 0,
        "Total_win": 0,
        "Total_loss": 0
    },
    "CH": {
        "Total_runs_scored": 0,
        "Total_runs_conceded": 0,
        "Total_overs_played": Decimal(0),
        "Total_Matches": 0,
        "Total_win": 0,
        "Total_loss": 0
    },
    "DD": {
        "Total_runs_scored": 0,
        "Total_runs_conceded": 0,
        "Total_overs_played": Decimal(0),
        "Total_Matches": 0,
        "Total_win": 0,
        "Total_loss": 0
    },
    "BB": {
        "Total_runs_scored": 0,
        "Total_runs_conceded": 0,
        "Total_overs_played": Decimal(0),
        "Total_Matches": 0,
        "Total_win": 0,
        "Total_loss": 0
    }
}

# Iterate over the existing sheet and update the values in the second sheet
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    max_overs = row[1]
    batting_1st_team = row[2]
    batting_2nd_team = row[3]
    batting_1st_team_total_runs = int(row[4])
    batting_1st_team_overs_played = Decimal(row[5])
    batting_1st_team_wickets_down = row[6]
    batting_2nd_team_total_runs = int(row[7])
    batting_2nd_team_overs_played = Decimal(row[8])
    batting_2nd_team_wickets_down = row[9]

    if batting_1st_team_total_runs > batting_2nd_team_total_runs:
        win = 1
        batting_1st_team_overs_played = max_overs  # Use the max_overs value instead of a fixed value
        batting_2nd_team_overs_played = max_overs  # Use the max_overs value instead of a fixed value
    else:
        win = 2
        batting_1st_team_overs_played = max_overs  # Use the max_overs value instead of a fixed value
        dot_index = str(batting_2nd_team_overs_played).find('.') #Here we can get the value like 13.2 we just find here . is therer
        if dot_index >= 0:
            batting_2nd_team_overs = Decimal(batting_2nd_team_overs_played[:dot_index]) + Decimal(batting_2nd_team_overs_played[dot_index:]) / 6
            batting_2nd_team_overs_played = batting_2nd_team_overs

    if batting_1st_team in teams:
        team_stats = teams[batting_1st_team]
        team_stats["Total_runs_scored"] += batting_1st_team_total_runs
        team_stats["Total_runs_conceded"] += batting_2nd_team_total_runs
        team_stats["Total_overs_played"] += batting_1st_team_overs_played + batting_2nd_team_overs_played
        team_stats["Total_Matches"] += 1

        if win == 1:
            team_stats["Total_win"] += 1
        else:
            team_stats["Total_loss"] += 1

    if batting_2nd_team in teams:
        team_stats = teams[batting_2nd_team]
        team_stats["Total_runs_scored"] += batting_2nd_team_total_runs
        team_stats["Total_runs_conceded"] += batting_1st_team_total_runs
        team_stats["Total_overs_played"] += batting_1st_team_overs_played + batting_2nd_team_overs_played
        team_stats["Total_Matches"] += 1

        if win == 2:
            team_stats["Total_win"] += 1
        else:
            team_stats["Total_loss"] += 1

# Calculate Net Run Rate (NRR) for each team
for team, data in teams.items():
    total_runs_scored = data["Total_runs_scored"]
    total_runs_conceded = data["Total_runs_conceded"]
    total_overs_played = data["Total_overs_played"]

    net_run_rate = (total_runs_scored - total_runs_conceded) / total_overs_played
    data["NRR"] = round(net_run_rate,4)

    # Find the row index of the team in the second sheet
    team_row_index = None
    for i, row in enumerate(points_worksheet.iter_rows(values_only=True)):
        if row[0] == team:
            team_row_index = i + 1  # Adding 2 to match the row index in the sheet (1 for header row)
            #print("team_row_index",team_row_index)
            break

    # Update the corresponding row in the second sheet or append a new row
    if team_row_index is not None:
        points_worksheet.cell(row=team_row_index, column=2).value = data["Total_Matches"]
        points_worksheet.cell(row=team_row_index, column=3).value = data["Total_win"]
        points_worksheet.cell(row=team_row_index, column=4).value = data["Total_loss"]
        points_worksheet.cell(row=team_row_index, column=5).value = data["Total_win"] * 2
        points_worksheet.cell(row=team_row_index, column=6).value = data["NRR"]
    else:
        new_row = [team, data["Total_Matches"], data["Total_win"], data["Total_loss"], data["Total_win"] * 2, data["NRR"]]
        points_worksheet.append(new_row)

# Save the changes to the workbook
workbook.save('C:\\cricket_scores.xlsx')
